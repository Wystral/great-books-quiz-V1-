"""
export_data.py

Reads the Great Books spreadsheet and exports structured JSON files into ./data/.

Output files:
  data/books.json           - All book records
  data/questions.json       - Quiz questions with metadata
  data/answer_options.json  - Answer options grouped by question_id
  data/scoring_rules.json   - Scoring rules per question/answer pair
  data/selection_rules.json - Selection rules grouped by rule_group
  data/genre_adjacency.json - Genre adjacency map

Run:
  python3 export_data.py
"""

import json
import os
import openpyxl

XLSX_PATH = os.path.join(
    os.path.dirname(__file__),
    "..",
    "attached_assets",
    "great_books_metadata_v29_tool_input_only_final_locked_vocab_c_1774247471031.xlsx",
)
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")


def load_workbook():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    return wb


def rows_to_dicts(ws):
    """Convert a worksheet to a list of dicts using the first row as headers."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        d = {}
        for h, v in zip(headers, row):
            d[h] = v
        result.append(d)
    return result


def slugify(text):
    """Create a stable URL-safe id from a book title."""
    import re
    s = str(text).lower().strip()
    s = re.sub(r"[^a-z0-9\s-]", "", s)
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")


def export_books(wb):
    """Export Book Metadata sheet, adding a stable slug id per book."""
    ws = wb["Book Metadata"]
    books = rows_to_dicts(ws)

    cleaned = []
    seen_ids = {}
    for b in books:
        themes_raw = b.get("themes") or ""
        themes = [t.strip() for t in str(themes_raw).split(";") if t.strip()] if themes_raw else []

        mood_raw = b.get("mood") or ""
        moods = [m.strip() for m in str(mood_raw).split(";") if m.strip()] if mood_raw else []

        base_id = slugify(b.get("work") or "unknown")
        if base_id in seen_ids:
            seen_ids[base_id] += 1
            book_id = f"{base_id}-{seen_ids[base_id]}"
        else:
            seen_ids[base_id] = 1
            book_id = base_id

        cleaned.append({
            "id": book_id,
            "work": b.get("work"),
            "author": b.get("author"),
            "era": b.get("era"),
            "era_section": b.get("era_section"),
            "genre": b.get("genre"),
            "themes": themes,
            "beginner_friendly": b.get("beginner_friendly"),
            "length": b.get("length"),
            "difficulty": b.get("difficulty"),
            "mood": moods,
            "author_blurb": b.get("author_blurb"),
            "work_synopsis": b.get("work_synopsis"),
            "why_read_this": b.get("why_read_this"),
        })

    return cleaned


def export_questions(wb):
    """Export Questions sheet."""
    ws = wb["Questions"]
    questions = rows_to_dicts(ws)
    result = []
    for q in questions:
        result.append({
            "display_order": q.get("display_order"),
            "question_id": q.get("question_id"),
            "question_text": q.get("question_text"),
            "primary_book_field": q.get("primary_book_field"),
            "max_choices": q.get("max_choices"),
            "response_type": q.get("response_type"),
            "scoring_note": q.get("scoring_note"),
        })
    result.sort(key=lambda x: x["display_order"] or 0)
    return result


def export_answer_options(wb):
    """Export Answer Options sheet, grouped by question_id."""
    ws = wb["Answer Options"]
    rows = rows_to_dicts(ws)

    grouped = {}
    for row in rows:
        qid = row.get("question_id")
        if qid not in grouped:
            grouped[qid] = []
        grouped[qid].append({
            "display_order": row.get("display_order"),
            "answer_option": row.get("answer_option"),
        })

    for qid in grouped:
        grouped[qid].sort(key=lambda x: x["display_order"] or 0)

    return grouped


def export_scoring_rules(wb):
    """Export Scoring Rules sheet."""
    ws = wb["Scoring Rules"]
    rows = rows_to_dicts(ws)

    rules = []
    for row in rows:
        rules.append({
            "rule_group": row.get("rule_group"),
            "question_id": row.get("question_id"),
            "answer_option": row.get("answer_option"),
            "book_field": row.get("book_field"),
            "book_value": row.get("book_value"),
            "score": row.get("score"),
            "notes": row.get("notes"),
        })
    return rules


def export_selection_rules(wb):
    """Export Selection Rules sheet, grouped by rule_group."""
    ws = wb["Selection Rules"]
    rows = rows_to_dicts(ws)

    grouped = {}
    for row in rows:
        rg = row.get("rule_group") or "ungrouped"
        if rg not in grouped:
            grouped[rg] = {}
        rule_name = row.get("rule_name")
        rule_value = row.get("rule_value")
        notes = row.get("notes")
        grouped[rg][rule_name] = {
            "value": rule_value,
            "notes": notes,
        }
    return grouped


def export_genre_adjacency(wb):
    """Export Genre Adjacency sheet as a map: selected_genre -> list of allowed genres."""
    ws = wb["Genre Adjacency"]
    rows = rows_to_dicts(ws)

    adjacency = {}
    for row in rows:
        sg = row.get("selected_genre")
        ag = row.get("allowed_book_genre")
        rel = row.get("relationship_note")
        if sg not in adjacency:
            adjacency[sg] = []
        adjacency[sg].append({
            "allowed_book_genre": ag,
            "relationship": rel,
        })
    return adjacency


def write_json(path, data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    size = os.path.getsize(path)
    print(f"  Wrote {path}  ({size:,} bytes)")


def main():
    print(f"Loading workbook from:\n  {XLSX_PATH}\n")
    wb = load_workbook()

    print("Exporting JSON files to ./data/ ...\n")

    books = export_books(wb)
    write_json(os.path.join(DATA_DIR, "books.json"), books)
    print(f"    => {len(books)} books exported\n")

    questions = export_questions(wb)
    write_json(os.path.join(DATA_DIR, "questions.json"), questions)
    print(f"    => {len(questions)} questions exported\n")

    answer_options = export_answer_options(wb)
    write_json(os.path.join(DATA_DIR, "answer_options.json"), answer_options)
    total_opts = sum(len(v) for v in answer_options.values())
    print(f"    => {total_opts} answer options across {len(answer_options)} questions\n")

    scoring_rules = export_scoring_rules(wb)
    write_json(os.path.join(DATA_DIR, "scoring_rules.json"), scoring_rules)
    print(f"    => {len(scoring_rules)} scoring rules exported\n")

    selection_rules = export_selection_rules(wb)
    write_json(os.path.join(DATA_DIR, "selection_rules.json"), selection_rules)
    print(f"    => {len(selection_rules)} rule groups exported\n")

    genre_adjacency = export_genre_adjacency(wb)
    write_json(os.path.join(DATA_DIR, "genre_adjacency.json"), genre_adjacency)
    print(f"    => {len(genre_adjacency)} genre entries exported\n")

    print("Stage 1 complete. File structure:")
    print()
    print("great-books-quiz/")
    print("├── export_data.py")
    print("└── data/")
    for fn in sorted(os.listdir(DATA_DIR)):
        size = os.path.getsize(os.path.join(DATA_DIR, fn))
        print(f"    ├── {fn}  ({size:,} bytes)")


if __name__ == "__main__":
    main()
